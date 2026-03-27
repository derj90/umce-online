#!/usr/bin/env python3
"""
Seed actividades_sdpa desde los xlsx de Salomé.
Genera SQL INSERT para ejecutar en Supabase DB.

Fuentes:
- registro_formacion_nov2025.xlsx (principal): 289 profs, actividades PIAC + formativas
- registro_formacion_online.xlsx (complementario): 94 profs, actividades formativas adicionales

Output: /tmp/seed-sdpa.sql
"""

import openpyxl
import re
import json
from datetime import datetime

# Activity → (linea, programa, tipo) mapping
ACTIVITY_MAP = {
    'Plan Instruccional de la Actividad Curricular PIAC': ('docencia', 'acompanamiento_focalizado', 'curso'),
    'Formacion para la tutoría virtual (15)': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'Formacion para la tutoría virtual': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'TIC en la enseñanza de la universidad (4)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'TIC en la enseñanza de la universidad': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Jornada de Formación en competencias digitales para la docencia UMCE (4)': ('integracion_tic', 'actualizacion_desarrollo', 'seminario'),
    'Jornada de Formación en competencias digitales para la docencia UMCE': ('integracion_tic', 'actualizacion_desarrollo', 'seminario'),
    'Taller chatGPT (2)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Taller chatGPT': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Conociendo la Competencia Genérica TIC (2)': ('integracion_tic', 'induccion', 'taller'),
    'Conociendo la Competencia Genérica TIC': ('integracion_tic', 'induccion', 'taller'),
    'Creación de Recursos Educativos con CANVA (2)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Creación de Recursos Educativos con CANVA': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Sácale Provecho a Zoom (2)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Sácale Provecho a Zoom': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Sácale provecho a zoom (2)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Inducción docente (2)': ('integracion_tic', 'induccion', 'induccion'),
    'Inducción docente': ('integracion_tic', 'induccion', 'induccion'),
    'Introducción a la IA y sus aplicaciones educativas (36)': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'Introducción a la IA y sus aplicaciones educativas': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'Dinamización de clases síncronas (4)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Dinamización de clases síncronas': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Creación de recursos interactivos para clases asíncronas (4)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Creación de recursos interactivos para clases asíncronas': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Elaboración de materiales didácticos Digitales para el aula (9)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Elaboración de materiales didácticos Digitales para el aula': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Google Workspace (27)': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'Google Workspace': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'Herramientas y Recursos para la Planificación de clases (4)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Herramientas y Recursos para la Planificación de clases': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Portafolios digitales (4)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Portafolios digitales': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Evaluación en línea (4)': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Evaluación en línea': ('integracion_tic', 'actualizacion_desarrollo', 'taller'),
    'Educación y Ciudadanía Digital (15)': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
    'Educación y Ciudadanía Digital': ('integracion_tic', 'actualizacion_desarrollo', 'curso'),
}

def extract_hours_from_name(name):
    """Extract hours from activity name like 'Taller chatGPT (2)' → 2"""
    m = re.search(r'\((\d+)\)\s*$', name.strip())
    return float(m.group(1)) if m else None

def clean_activity_name(name):
    """Normalize activity name"""
    return name.strip().rstrip('\n').strip()

def get_activity_info(name):
    """Map activity name to (linea, programa, tipo)"""
    cleaned = clean_activity_name(name)
    if cleaned in ACTIVITY_MAP:
        return ACTIVITY_MAP[cleaned]
    # Try without hours suffix
    base = re.sub(r'\s*\(\d+\)\s*$', '', cleaned)
    if base in ACTIVITY_MAP:
        return ACTIVITY_MAP[base]
    # Default: integracion_tic
    return ('integracion_tic', 'actualizacion_desarrollo', 'otro')

def escape_sql(val):
    """Escape string for SQL"""
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def parse_nov2025():
    """Parse main file: each row may have 1 activity"""
    records = []
    wb = openpyxl.load_workbook('/Users/coordinacion/Downloads/registro_formacion_nov2025.xlsx', read_only=True)
    ws = wb['Sólo profes']

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue
        name, rut, email, dept, fac, prog, ac_piac, finished, activity, hours = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]

        if not name or not activity:
            continue

        # Normalize email
        email = str(email).strip().lower() if email else None
        if not email:
            continue

        act_name = clean_activity_name(str(activity))
        linea, programa, tipo = get_activity_info(act_name)

        hrs = float(hours) if hours else extract_hours_from_name(act_name)
        if not hrs:
            continue

        records.append({
            'docente_email': email,
            'nombre_actividad': act_name,
            'descripcion': f'Programa: {prog}. AC PIAC: {ac_piac}' if prog else None,
            'linea': linea,
            'programa': programa,
            'tipo': tipo,
            'horas_cronologicas': hrs,
            'estado': 'completada',
            'plataforma': 'moodle_evirtual',
            'registrado_por': 'sistema_import_xlsx',
            'notas': json.dumps({'fuente': 'registro_formacion_nov2025.xlsx', 'nombre': str(name), 'rut': str(rut), 'departamento': str(dept), 'facultad': str(fac)}, ensure_ascii=False),
        })

    return records

def parse_online():
    """Parse complementary file: multiline activities per professor"""
    records = []
    wb = openpyxl.load_workbook('/Users/coordinacion/Downloads/registro_formacion_online.xlsx', read_only=True)
    ws = wb['Sólo profes']

    for row in ws.iter_rows(min_row=3, values_only=True):
        name, rut, email, dept, fac, prog, ac_piac, finished, formation, activities, hours = (
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10]
        )

        if not name or not activities:
            continue

        email = str(email).strip().lower() if email else None
        if not email:
            continue

        # Split multiline activities
        activity_lines = [a.strip() for a in str(activities).split('\n') if a.strip()]

        for act_line in activity_lines:
            act_name = clean_activity_name(act_line)
            linea, programa, tipo = get_activity_info(act_name)
            hrs = extract_hours_from_name(act_name)

            if not hrs:
                # Use total hours divided by activity count as fallback
                if hours and len(activity_lines) > 0:
                    hrs = float(hours) / len(activity_lines)
                else:
                    continue

            records.append({
                'docente_email': email,
                'nombre_actividad': act_name,
                'descripcion': f'Programa: {prog}. AC PIAC: {ac_piac}' if prog else None,
                'linea': linea,
                'programa': programa,
                'tipo': tipo,
                'horas_cronologicas': hrs,
                'estado': 'completada' if str(finished).lower() in ('si', 'sí', '1', '1.0') else 'en_progreso',
                'plataforma': 'moodle_evirtual',
                'registrado_por': 'sistema_import_xlsx',
                'notas': json.dumps({'fuente': 'registro_formacion_online.xlsx', 'nombre': str(name), 'rut': str(rut), 'departamento': str(dept), 'facultad': str(fac)}, ensure_ascii=False),
            })

    return records

def deduplicate(records):
    """Deduplicate by email + activity name"""
    seen = set()
    unique = []
    for r in records:
        key = (r['docente_email'], r['nombre_actividad'])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique

def generate_sql(records):
    """Generate SQL INSERT statements"""
    lines = [
        '-- Generated by seed-sdpa.py on ' + datetime.now().isoformat(),
        '-- Records: ' + str(len(records)),
        '',
        'BEGIN;',
        '',
    ]

    for r in records:
        lines.append(f"""INSERT INTO portal.actividades_sdpa
    (docente_email, nombre_actividad, descripcion, linea, programa, tipo,
     horas_cronologicas, estado, plataforma, registrado_por, notas, fecha_inicio)
VALUES
    ({escape_sql(r['docente_email'])}, {escape_sql(r['nombre_actividad'])},
     {escape_sql(r['descripcion'])}, {escape_sql(r['linea'])},
     {escape_sql(r['programa'])}, {escape_sql(r['tipo'])},
     {r['horas_cronologicas']}, {escape_sql(r['estado'])},
     {escape_sql(r['plataforma'])}, {escape_sql(r['registrado_por'])},
     {escape_sql(r['notas'])}, '2025-01-01')
ON CONFLICT DO NOTHING;""")
        lines.append('')

    # Generate progreso_certificaciones based on accumulated TIC hours
    lines.append('-- ============================================')
    lines.append('-- Progreso certificaciones TIC (auto-calculated)')
    lines.append('-- ============================================')
    lines.append('')

    # Accumulate TIC hours per email
    tic_hours = {}
    for r in records:
        if r['linea'] == 'integracion_tic' and r['estado'] == 'completada':
            email = r['docente_email']
            tic_hours[email] = tic_hours.get(email, 0) + r['horas_cronologicas']

    # Cert thresholds: inicial=27h, intermedio=54h, avanzado=81h
    cert_slugs = [
        ('tic_inicial', 27),
        ('tic_intermedio', 54),
        ('tic_avanzado', 81),
    ]

    for email, hours in sorted(tic_hours.items()):
        for slug, threshold in cert_slugs:
            if hours >= threshold:
                estado = 'requisitos_cumplidos'
            elif hours > 0:
                estado = 'en_progreso'
            else:
                continue

            lines.append(f"""INSERT INTO portal.progreso_certificaciones
    (docente_email, certificacion_id, horas_acumuladas, estado)
VALUES
    ({escape_sql(email)},
     (SELECT id FROM portal.certificaciones_sdpa WHERE slug = {escape_sql(slug)}),
     {min(hours, threshold * 1.5)}, {escape_sql(estado)})
ON CONFLICT (docente_email, certificacion_id) DO UPDATE SET
    horas_acumuladas = EXCLUDED.horas_acumuladas,
    estado = EXCLUDED.estado,
    updated_at = now();""")
            lines.append('')

            # Only create entries for levels they've reached or are working toward
            if hours < threshold:
                break

    lines.append('COMMIT;')
    return '\n'.join(lines)

if __name__ == '__main__':
    print('Parsing nov2025...')
    r1 = parse_nov2025()
    print(f'  {len(r1)} records')

    print('Parsing online...')
    r2 = parse_online()
    print(f'  {len(r2)} records')

    all_records = r1 + r2
    print(f'Total before dedup: {len(all_records)}')

    unique = deduplicate(all_records)
    print(f'Total after dedup: {len(unique)}')

    # Stats
    emails = set(r['docente_email'] for r in unique)
    print(f'Unique professors: {len(emails)}')

    tic_hours = {}
    for r in unique:
        if r['linea'] == 'integracion_tic' and r['estado'] == 'completada':
            tic_hours[r['docente_email']] = tic_hours.get(r['docente_email'], 0) + r['horas_cronologicas']

    bajo_inicial = sum(1 for h in tic_hours.values() if h < 27)
    inicial = sum(1 for h in tic_hours.values() if 27 <= h < 54)
    intermedio = sum(1 for h in tic_hours.values() if 54 <= h < 81)
    avanzado = sum(1 for h in tic_hours.values() if h >= 81)
    sin_tic = len(emails) - len(tic_hours)

    print(f'\nCertificación TIC:')
    print(f'  Sin horas TIC: {sin_tic}')
    print(f'  Bajo inicial (<27h): {bajo_inicial}')
    print(f'  Inicial (27-53h): {inicial}')
    print(f'  Intermedio (54-80h): {intermedio}')
    print(f'  Avanzado (81h+): {avanzado}')

    sql = generate_sql(unique)
    with open('/tmp/seed-sdpa.sql', 'w') as f:
        f.write(sql)
    print(f'\nSQL written to /tmp/seed-sdpa.sql ({len(sql)} bytes)')
